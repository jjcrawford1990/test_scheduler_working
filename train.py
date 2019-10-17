import openpyxl
from datetime import *

trains_dates_excel = 'C:\\Users\\Josh\\Desktop\\trains&dates.xlsx'  # path of the start dates & trains file to be read
schedule_excel = 'C:\\Users\\Josh\\Desktop\\schedule.xlsx'  # path of the start dates & trains file to be read



class LiveSchedule:
    test_hours_list=[]
    test_hours_total = 0
    no_shifts = 0
    shift_length = 7.5
    remaining_shifts = 0


    @classmethod
    def scheduleHours(cls):
        cls.schedule_wb = openpyxl.load_workbook(schedule_excel)  # create a class object of the active workbook within trains_dates excel
        cls.schedule_active_wb = cls.schedule_wb.active  # create a class object of the active workbook
        cls.schedule_ws = cls.schedule_wb.worksheets[0]  # create a class object of the first worksheet within trains_dates excel
        cls.no_tests = cls.schedule_ws.max_row - 1
        for i in range(cls.no_tests):
            cls.test_hours_list.append(cls.schedule_active_wb.cell(row=i+2, column=3).value)
        for i in range(cls.no_tests):
            cls.test_hours_total += cls.test_hours_list[i]
        cls.no_shifts = round(cls.test_hours_total / cls.shift_length) #rounded number of shifts to complete testing

    def remainingHours(self):
        self.test_list = []
        self.test_completion = []
        self.test_area_id = []
        self.actual_schedule = {'A': 0, 'B': LiveSchedule.no_shifts, 'C': 0}
        for i in range(LiveSchedule.no_tests):
            train_number_worksheet = 'T' + str(self.unit_number)
            self.test_list.append(LiveSchedule.schedule_active_wb.cell(row=i+2, column=1).value)
            self.test_completion.append(LiveSchedule.schedule_wb[train_number_worksheet].cell(row=i+2, column=6).value) #open worksheet of train schedule
            self.test_area_id.append(LiveSchedule.schedule_active_wb.cell(row=i+2, column=7).value)
        self.zipped_tests = zip(self.test_list, LiveSchedule.test_hours_list, self.test_completion, self.test_area_id)
        remaining_hours_subtraction = 0
        for w, x, y, z in self.zipped_tests:
            if y == 'Y':
                remaining_hours_subtraction += x
                self.actual_schedule[z] - (x/LiveSchedule.shift_length)
            else:
                continue
        self.remaining_hours = LiveSchedule.test_hours_total - remaining_hours_subtraction
        self.remaining_shifts = round(self.remaining_hours / LiveSchedule.shift_length)
        self.live_schedule = {}
        temp_sched_list = ['B'] * self.remaining_shifts
        for i in range(self.remaining_shifts):  # for the number of task types(this could be days), iterate over
            delta_date = date.today() + timedelta(days=i)
            weekday_of_date = date.isoweekday(delta_date)  # weekday (1-7) of current date
            if weekday_of_date <= 5:  # if weekday, add to dictionary
                self.live_schedule[delta_date] = temp_sched_list[i]  # in schededule dictionary, key is day, element is task ID
            else:
                # else, iterate date by 1 until a weekday less than 5 (monday, 1) is reached
                continue

class Train(LiveSchedule):

    no_trains = 0 #class attribute for number of trains
    obj_generation_count = 0 #to count if this is initial or re-generation of train objects

    @classmethod
    def trainCount(cls):
        Train.trains_dates_wb = openpyxl.load_workbook(trains_dates_excel)  # create a class object of the active workbook within trains_dates excel
        Train.trains_dates_active_wb = Train.trains_dates_wb.active  # create a class object of the active workbook
        Train.trains_dates_ws = Train.trains_dates_wb.worksheets[0]  # create a class object of the first worksheet within trains_dates excel
        cls.no_trains = Train.trains_dates_ws.max_row - 1  # count the rows and minus the header row


    def assignTrainData(self, train_index):
        self.ind = train_index #attribute which is index number of train object within the 'train' list
        unit_number_cell = Train.trains_dates_active_wb.cell(row=self.ind + 2, column=1)  # iterate through the 1st row to assign train numbers
        start_date_cell = Train.trains_dates_active_wb.cell(row=self.ind + 2, column=2)  # iterate through the 1st row to assign train numbers
        project_cell = Train.trains_dates_active_wb.cell(row=self.ind + 2, column=3)  # iterate through the 1st row to assign train numbers
        schedule_id_cell = Train.trains_dates_active_wb.cell(row=self.ind + 2, column=4)  # iterate through the 1st row to assign train numbers
        self.unit_number = unit_number_cell.value #assign unit number attribute
        self.start_date = datetime.strptime(start_date_cell.value, '%Y %m %d').date()  # assign datetime attribute from start date cell value
        self.project = project_cell.value  # assign project attribute
        self.schedule_id = schedule_id_cell.value  # assign schedule id attribute

    def objGenerate(): #can re-run this method to re-generate the train list, for example if another module/application can edit the excel?
        try:
            train[1]
        except:
           print('List is Yet to exist')
           for i in range(Train.no_trains): #first object already created
                train[i].assignTrainData(i)  # pass i (index number) as an argument to assignTrainData method
                train[i].scheduleGenerate()
                train.append(Train())
        else:
            print('List Exists, now we destroy it!')
            Train.trainCount() #recount no. trains
            if Train.no_trains < (len(train)):
            # delete list element namespaces above current index e.g. current list is 10, new is 9. delete elements 10-> last element inclusive
                del (train[Train.no_trains-1:len(train)-1])
                for i in range(Train.no_trains):
                    train[i].assignTrainData(i)  # pass i (index number) as an argument to assignTrainData method

    def taskTypes():
        Train.task_types_ws = Train.trains_dates_wb.worksheets[1]  # object of task types worksheet
        Train.no_of_task_types = Train.task_types_ws.max_row - 1  # iterate through the 1st row to assign train numbers
        Train.task_types = {} #task type dictionary instantiate
        for i in range(Train.no_of_task_types):
            #the below adds key (Letter identifier) and element (Descriptor) to dictionary
            Train.task_types[Train.task_types_ws.cell(row=i + 2, column=1).value] = Train.task_types_ws.cell(row=i + 2, column=2).value

    def scheduleDictionary():
        Train.schedules_ws = Train.trains_dates_wb.worksheets[2]  # object of schedules worksheet
        Train.no_of_schedules = Train.schedules_ws.max_row - 1  # iterate through the 1st row to assign train numbers
        Train.schedules = {}  # schedule dictionary instantiate
        for i in range(Train.no_of_schedules):
            # the below adds key (schedule type) and element (construction) to dictionary
            Train.schedules[Train.schedules_ws.cell(row=i + 2, column=1).value] = Train.schedules_ws.cell(row=i + 2, column=2).value

    def scheduleGenerate(self):
        self.schedule = {} #create empty dictionary. this will store the objects schedule
        sched_list = Train.schedules.get(self.schedule_id) #create list & get schedule dict element with our schedule_id as the key
        self.split_sched_list = sched_list.split(',') #split the elements into a list, elements delimited by a comma
        temp_sched_list = [] #for storing a list of the ID's
        for i in self.split_sched_list:
            sliced_id = i[0:1] #task type
            sliced_no_days = i[1:] #no of days assigned
            for x in range(int(sliced_no_days)):
                temp_sched_list.append(sliced_id)
        for i in range(len(temp_sched_list)): #for the number of task types(this could be days), iterate over
            delta_date = self.start_date + timedelta(days=i)
            weekday_of_date = date.isoweekday(delta_date)  # weekday (1-7) of current date
            if weekday_of_date <= 5: #if weekday, add to dictionary
                self.schedule[delta_date] = temp_sched_list[i] #in schededule dictionary, key is day, element is task ID
            else:
                #else, iterate date by 1 until a weekday less than 5 (monday, 1) is reached
                continue


    def createTrain():
        new_unit_number = 155 #new number getter method from tkinter frame
        new_start_date = 150110 #new start date getter method from tkinter frame
        new_project = 'MILANO' #new project getter method from tkinter frame
        Train.trains_dates_ws['A15'] = new_unit_number #write the new value
        Train.trains_dates_ws['B15'] = new_start_date #write the new value
        Train.trains_dates_ws['C15'] = new_project #write the new value
        Train.trains_dates_wb.save(trains_dates_excel) #save the workbook
        #train.append(Train())  #add new train object to list
        #Train.no_trains +=1 #increment number of trains active
        #train[Train.no_trains-1].assignTrainData(Train.no_trains-1) #assign data to new train object

    def deleteTrain():
        pass

LiveSchedule.scheduleHours()
Train.trainCount()
Train.taskTypes()
Train.scheduleDictionary()
print(Train.no_trains, 'Trains Exist')
train = [] #create empty train list
train.append(Train()) #create train list object for first index only
Train.objGenerate() #run methods and generate all other train objects
for i in range(Train.no_trains):
    train[i].remainingHours()